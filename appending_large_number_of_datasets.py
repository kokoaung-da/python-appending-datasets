import warnings
warnings.filterwarnings("ignore")

import os
import pandas as pd
import re
import math
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# --- Config ---
# Please adjust these paths to match your local machine
folder_path = r"input folder path"
output_dir = r"output folder path"
os.makedirs(output_dir, exist_ok=True)

# Standard reference column names (normalized)
standard_columns = [
    "year", "month", "erssn", "ername", "eessn", "eename", "minc",
    "ss1eerate", "ss1errate", "ss1eeconamt", "ss1erconamt",
    "ss2eerate", "ss2errate", "ss2eeconamt", "ss2erconamt", "totalconamt"
]
STANDARD_COLUMN_COUNT = len(standard_columns)

# --- Helpers ---
def normalize_string(text):
    """
    Normalizes a string by removing newlines from wrap text,
    stripping all non-alphanumeric characters, converting to lowercase,
    and removing leading/trailing whitespace.
    This is used for both sheet names and column headers.
    """
    if not isinstance(text, str):
        text = str(text)
    # Replace newline characters with a space
    clean_text = text.replace('\n', ' ').replace('\r', ' ')
    # Remove all non-alphanumeric characters
    clean_text = re.sub(r'[\W_]+', '', clean_text)
    # Convert to lowercase and strip whitespace
    return clean_text.lower().strip()

def normalize_headers(cols):
    """Normalizes a list of column headers using the helper function."""
    return [normalize_string(col) for col in cols]

# --- Main Logic ---
combined_data = []
mismatched_files = []
failed_files = []

# Get a list of all Excel and CSV files, including temporary ones
all_files = [
    os.path.join(root, file)
    for root, _, files in os.walk(folder_path)
    for file in files if (file.endswith('.xlsx') or file.endswith('.csv'))
]

# --- NEW: Sort the files to ensure a consistent processing order ---
all_files.sort()

print(f"🚀 Starting process for {len(all_files)} files...")

for file_path in all_files:
    try:
        df = None
        # --- Handle both CSV and XLSX files ---
        if file_path.lower().endswith('.csv'):
            # For CSV files, read them directly
            df = pd.read_csv(file_path)
        
        elif file_path.lower().endswith('.xlsx'):
            # For Excel files, use the strict sheet detection logic
            sheet_to_read = None
            workbook = load_workbook(file_path, read_only=True)
            visible_sheet_names = [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == 'visible']
            
            for name in visible_sheet_names:
                if normalize_string(name) == 'sheet1':
                    sheet_to_read = name
                    break
            
            if sheet_to_read is None:
                raise Exception("No VISIBLE sheet named 'sheet1' (after normalization) was found.")
            
            df = pd.read_excel(file_path, sheet_name=sheet_to_read)

        # Skip empty rows
        df.dropna(how='all', inplace=True)

        if df.empty:
            print(f"INFO: File '{os.path.basename(file_path)}' is empty after removing blank rows. Skipping.")
            continue

        # --- Column & Data Processing ---
        original_cols = df.columns.tolist()
        normalized_cols = normalize_headers(original_cols)
        
        processed_df = None

        if set(normalized_cols) == set(standard_columns):
            rename_dict = {orig_col: std_col for orig_col in original_cols for std_col in standard_columns if normalize_string(orig_col) == std_col}
            df.rename(columns=rename_dict, inplace=True)
            processed_df = df[standard_columns]
        elif len(normalized_cols) == STANDARD_COLUMN_COUNT:
            print(f"WARN: Headers in '{os.path.basename(file_path)}' don't match, but column count is correct. Renaming columns by position.")
            df.columns = standard_columns
            processed_df = df
        else:
            mismatched_files.append({
                'file': os.path.basename(file_path),
                'directory': os.path.dirname(file_path),
                'missing': list(set(standard_columns) - set(normalized_cols)),
                'extra': list(set(normalized_cols) - set(standard_columns)),
                'found_count': len(normalized_cols)
            })
            continue

        if processed_df is not None:
            processed_df['source_directory'] = os.path.dirname(file_path)
            processed_df['source_filename'] = os.path.basename(file_path)
            combined_data.append(processed_df)

    except Exception as e:
        failed_files.append({
            'file': os.path.basename(file_path),
            'directory': os.path.dirname(file_path),
            'reason': str(e)
        })

# --- Save Combined Data ---
output_paths_str = ""
if combined_data:
    final_df = pd.concat(combined_data, ignore_index=True)
    
    if len(final_df) > 500000:
        chunk_size = 500000
        num_files = math.ceil(len(final_df) / chunk_size)
        output_paths = []

        print(f"\nCombined data has {len(final_df)} rows. Splitting into {num_files} files...")

        for i in range(num_files):
            start_row = i * chunk_size
            end_row = (i + 1) * chunk_size
            chunk_df = final_df.iloc[start_row:end_row]
            
            file_path = os.path.join(output_dir, f'all_combined_data_{i+1}.xlsx')
            output_paths.append(file_path)
            
            print(f"Saving part {i+1} ({len(chunk_df)} rows) to {file_path}...")
            chunk_df.to_excel(file_path, index=False)
        
        output_paths_str = ", ".join(output_paths)
        print(f"\nSplitting complete.")
    else:
        single_path = os.path.join(output_dir, 'all_combined_data.xlsx')
        final_df.to_excel(single_path, index=False)
        output_paths_str = single_path
        print(f"\nCombined data has {len(final_df)} rows, saved as a single XLSX file.")

# --- Save Reports without Hyperlinks ---
if mismatched_files:
    wb = Workbook()
    ws = wb.active
    ws.title = "MismatchedFiles"
    ws.append(["Sr. No.", "File Name", "Directory", "Expected Column Count", "Found Column Count", "Missing Standard Columns", "Extra Columns Found"])
    for idx, item in enumerate(mismatched_files, start=1):
        ws.append([
            idx, 
            item['file'], 
            item['directory'], 
            STANDARD_COLUMN_COUNT, 
            item['found_count'], 
            str(item['missing']), 
            str(item['extra'])
        ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(os.path.join(output_dir, "mismatched_files_report.xlsx"))

if failed_files:
    wb = Workbook()
    ws = wb.active
    ws.title = "FailedFiles"
    ws.append(["Sr. No.", "File Name", "Directory", "Reason of Failed"])
    for idx, item in enumerate(failed_files, start=1):
        ws.append([
            idx, 
            item['file'], 
            item['directory'], 
            item['reason']
        ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    wb.save(os.path.join(output_dir, "failed_files.xlsx"))

# --- Final Summary ---
print("\n--- 📊 Process Summary ---")
print(f"✅ Successfully combined files: {len(combined_data)}")
print(f"⚠️ Files with mismatched columns: {len(mismatched_files)}")
print(f"❌ Failed to read files: {len(failed_files)}")
print("-" * 25)
if output_paths_str:
    print(f"📁 Combined data saved to: {output_paths_str}")
if mismatched_files:
    print(f"⚠️ Mismatched report saved to: {os.path.join(output_dir, 'mismatched_files_report.xlsx')}")
if failed_files:
    print(f"❌ Failure report saved to: {os.path.join(output_dir, 'failed_files.xlsx')}")
print("--- Process Complete ---")
